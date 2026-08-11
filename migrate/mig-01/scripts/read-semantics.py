# -*- coding: utf-8 -*-
"""MIG-01 试验 B 辅助：读取单个 xlsx 并输出 openpyxl 语义快照（stdout JSON）。"""
import json
import sys

from openpyxl import load_workbook


def cell_semantics(cell):
    value = cell.value
    if value is None:
        return {'kind': 'empty'}
    if isinstance(value, str):
        kind = 's'
    elif isinstance(value, (int, float)):
        kind = 'n'
    elif isinstance(value, bool):
        kind = 'b'
    elif hasattr(value, 'isoformat'):
        kind = 'd'
        value = value.isoformat()[:10]
    else:
        kind = 's'
        value = str(value)
    return {'kind': kind, 'v': value, 'bold': bool(cell.font and cell.font.bold),
            'fmt': '' if (cell.number_format or '') == 'General' else (cell.number_format or '')}


def main(path):
    wb = load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        grid = [[cell_semantics(cell) for cell in row] for row in ws.iter_rows()]
        sheets.append({
            'name': ws.title,
            'dims': ws.calculate_dimension(),
            'freeze': str(ws.freeze_panes) if ws.freeze_panes else '',
            'merges': sorted(str(r) for r in ws.merged_cells.ranges),
            'widths': {col: round((d.width or 8.43), 2) for col, d in
                       sorted(ws.column_dimensions.items()) if not d.hidden},
            'grid': grid,
        })
    wb.close()
    print(json.dumps({'sheets': sheets}, ensure_ascii=False))
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv[1]))
