# -*- coding: utf-8 -*-
"""MIG-01 试验 B：生成黄金工作簿与 openpyxl 语义快照。

用后端导出接口（TestClient + 临时库 + 固定夹具）生成各类导出文件，
再用 openpyxl 提取语义快照（sheet 名/合并/冻结/列宽/单元格值/日期/公式/加粗），
供 exceljs 端对比。输出 migrate/mig-01/out/excel/。
"""
import json
import os
import sys

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'backend'))
sys.path.insert(0, PROJECT_ROOT)
os.environ['WORKBENCH_BUSINESS_DATE'] = '2026-04-15'
os.environ['WORKBENCH_KB_DIR'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'out', 'excel-kb')

from app import db  # noqa: E402
from backend.tests.helpers import enroll_all_students  # noqa: E402

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'out', 'excel')
os.makedirs(OUT, exist_ok=True)


def cell_semantics(cell):
    value = cell.value
    kind = 'empty'
    if value is None:
        return {'kind': 'empty'}
    if isinstance(value, str):
        kind = 's'
    elif isinstance(value, (int, float)):
        kind = 'n'
    elif isinstance(value, bool):
        kind = 'b'
    elif hasattr(value, 'isoformat'):
        kind = 'd'
        value = value.isoformat()[:10]
    elif value.startswith('='):
        kind = 'f'
    return {'kind': kind, 'v': value, 'bold': bool(cell.font and cell.font.bold),
            'fmt': '' if (cell.number_format or '') == 'General' else (cell.number_format or '')}


def workbook_semantics(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        grid = []
        for row in ws.iter_rows():
            grid.append([cell_semantics(cell) for cell in row])
        sheets.append({
            'name': ws.title,
            'dims': ws.calculate_dimension(),
            'freeze': str(ws.freeze_panes) if ws.freeze_panes else '',
            'merges': sorted(str(r) for r in ws.merged_cells.ranges),
            'widths': {col: round((d.width or 8.43), 2) for col, d in
                       sorted(ws.column_dimensions.items()) if not d.hidden},
            'grid': grid,
        })
    wb.close()
    return {'sheets': sheets}


def main():
    from fastapi.testclient import TestClient
    from app import app as application
    import shutil

    tmp = os.path.join(OUT, '.tmp')
    shutil.rmtree(tmp, ignore_errors=True)
    os.makedirs(tmp, exist_ok=True)
    old = (db.DB_PATH, db.DATA_DIR)
    db.close()
    db.DATA_DIR = tmp
    db.DB_PATH = os.path.join(tmp, 'workbench.db')
    conn = db.get_conn()
    fixture = json.load(open(os.path.join(
        PROJECT_ROOT, 'backend', 'tests', 'fixtures', 'p0_demo.json'), encoding='utf-8'))
    for student in fixture['students']:
        conn.execute('INSERT INTO students(学号, 姓名, 性别, 班级任职) VALUES(?,?,?,?)',
                     tuple(student.get(c, '') for c in ('学号', '姓名', '性别', '班级任职')))
    conn.commit()
    enroll_all_students()

    client = TestClient(application, raise_server_exceptions=False)
    headers = {'x-workbench-class': '1', 'x-workbench-term': '1'}
    # 准备导出所需的业务数据
    client.post('/api/attendance/daily', json={
        'date': '2026-04-15', 'scene': '常规到校',
        'records': [{'student_id': 1, 'status': '出勤'}]}, headers=headers)
    client.post('/api/comments/entries', json={
        'student_id': 1, 'comment_type': '学期评语', 'content': '黄金导出评语'}, headers=headers)
    client.post('/api/fund/entries', json={
        'direction': '支出', 'amount': 42.5, 'description': '黄金导出支出'}, headers=headers)
    client.post('/api/score-config/subjects', json={'name': '物理', 'full_score': 100, 'type': '选考'},
                headers=headers)
    client.post('/api/score-config/exams', json={'name': '月考1', 'exam_date': '2026-04-20'},
                headers=headers)
    client.post('/api/exams/import/commit', json={
        'filename': 'golden.xlsx', 'request_id': 'golden-excel-scores',
        'rows': [{'row': 1, 'valid': True, 'student_id': 1, 'exam_name': '月考1',
                  'exam_date': '2026-04-20', 'subject': '物理', 'score': 92.0, 'rank': 1,
                  'record_status': '正常', 'note': ''}]}, headers=headers)

    targets = [
        ('students-template', 'GET', '/api/students/template', {}),
        ('students-export', 'GET', '/api/students/export', {}),
        ('sheet-export', 'GET', '/api/export/sheet/学生信息总表', {}),
        ('attendance-report', 'GET', '/api/export/report/attendance',
         {'date_from': '2026-04-01', 'date_to': '2026-04-30'}),
        ('scores-report', 'GET', '/api/export/report/scores', {}),
        ('comments-export', 'GET', '/api/export/sheet/评语管理', {}),
        ('fund-export', 'GET', '/api/export/sheet/班费管理', {}),
        ('health-export', 'GET', '/api/health/summary/export', {}),
    ]
    summary = {}
    for name, method, path, params in targets:
        print(f'导出 {name} ...', flush=True)
        response = client.request(method, path, params=params, headers=headers)
        if response.status_code != 200:
            summary[name] = {'status': response.status_code, 'error': response.text[:200]}
            continue
        file_path = os.path.join(OUT, f'{name}.xlsx')
        with open(file_path, 'wb') as target:
            target.write(response.content)
        semantics = workbook_semantics(file_path)
        with open(os.path.join(OUT, f'{name}.openpyxl.json'), 'w', encoding='utf-8') as target:
            json.dump(semantics, target, ensure_ascii=False, indent=1)
        summary[name] = {'status': 200, 'sheets': [s['name'] for s in semantics['sheets']],
                         'bytes': len(response.content)}

    db.close()
    db.DB_PATH, db.DATA_DIR = old
    shutil.rmtree(tmp, ignore_errors=True)
    with open(os.path.join(OUT, 'summary.json'), 'w', encoding='utf-8') as target:
        json.dump(summary, target, ensure_ascii=False, indent=1)
    print(json.dumps(summary, ensure_ascii=False, indent=1))
    return 0


if __name__ == '__main__':
    sys.exit(main())
