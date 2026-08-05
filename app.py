# -*- coding: utf-8 -*-
"""
美美大王工作台 v2.1 - Flask 后端
修复：公式列手动计算、班费余额数值化、导出功能
"""
from flask import Flask, jsonify, request, send_from_directory, send_file
from openpyxl import load_workbook
from datetime import datetime
import os
import re
import tempfile

app = Flask(__name__, static_folder='static', static_url_path='')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_BANZHUREN = os.path.join(BASE_DIR, '班主任工作台', '班主任工作台.xlsx')
XLSX_HEALTH = os.path.join(BASE_DIR, '健康管理', '健康追踪表.xlsx')
KB_DIR = os.path.join(BASE_DIR, '知识库')

BZR_SHEETS = [
    '学生信息总表', '特殊学生档案', '考勤管理', '成绩跟踪',
    '日常行为积分', '家校沟通记录', '谈心记录', '班会记录',
    '班主任日志', '班费管理', '座位表', '评语管理',
    '班委管理', '班级活动', '工作计划总结'
]
HEALTH_SHEETS = ['体重体脂追踪', '运动记录', '睡眠记录', '饮食记录', '月度总结']

SHEET_FILE = {}
for s in BZR_SHEETS:
    SHEET_FILE[s] = XLSX_BANZHUREN
for s in HEALTH_SHEETS:
    SHEET_FILE[s] = XLSX_HEALTH


# ==================== 工具函数 ====================

def _is_formula(val):
    """判断值是否为公式字符串"""
    return isinstance(val, str) and val.startswith('=')


def _safe_number(ws, row, col):
    """
    安全读取数值：如果是公式则返回 None（由调用方手动计算），
    如果是数值则返回数值，否则返回 None。
    """
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if _is_formula(val):
        return None  # 标记为需要手动计算
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _sum_range(ws, row, start_col, end_col):
    """手动计算一行中某范围内数值的总和"""
    total = 0
    has_any = False
    for c in range(start_col, end_col + 1):
        val = ws.cell(row=row, column=c).value
        if val is not None and not _is_formula(val):
            try:
                total += float(val)
                has_any = True
            except (ValueError, TypeError):
                pass
    return total if has_any else None


def _get_wb(sheet_name):
    """获取工作表对应的工作簿"""
    fp = SHEET_FILE.get(sheet_name)
    if not fp:
        return None, None
    wb = load_workbook(fp, data_only=False)
    ws = wb[sheet_name]
    return wb, ws


def _read_sheet(ws, header_row=3):
    """读取工作表数据，公式列尝试显示为数值（无法计算时显示为空）"""
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        headers.append(str(val).replace('\n', '') if val is not None else '')

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_data = []
        has_data = False
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                row_data.append(None)
            elif _is_formula(val):
                # 公式列 → 尝试手动计算（简单公式）
                computed = _try_eval_formula(ws, r, c, val)
                row_data.append(computed)
                if computed is not None:
                    has_data = True
            else:
                row_data.append(val)
                has_data = True
        if has_data:
            rows.append(row_data)

    return {'headers': headers, 'rows': rows, 'data_start_row': header_row + 1}


def _try_eval_formula(ws, row, col, formula_str):
    """尝试解析简单公式，返回计算结果或 None"""
    f = formula_str.strip()
    # =SUM(C4:H4) 模式
    m = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', f, re.IGNORECASE)
    if m:
        c1 = _col_to_num(m.group(1))
        r1 = int(m.group(2))
        c2 = _col_to_num(m.group(3))
        r2 = int(m.group(4))
        if r1 == r2 == row:
            return _sum_range(ws, row, c1, c2)
        return None

    # =IF(C4="","",...) 模式 → 返回 None，让调用方手动处理
    if f.upper().startswith('=IF('):
        return None

    # =COUNT(...) 模式
    m = re.match(r'^=COUNT\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', f, re.IGNORECASE)
    if m:
        c1 = _col_to_num(m.group(1))
        r1 = int(m.group(2))
        c2 = _col_to_num(m.group(3))
        r2 = int(m.group(4))
        if r1 == r2 == row:
            count = 0
            for c in range(c1, c2 + 1):
                val = ws.cell(row=row, column=c).value
                if val is not None:
                    count += 1
            return count
        return None

    return None


def _col_to_num(col_str):
    """A→1, B→2, ... Z→26, AA→27"""
    n = 0
    for ch in col_str.upper():
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def _calc_balance(ws):
    """手动计算班费余额（从第4行开始遍历 B/C 列）"""
    balance = 0
    for r in range(4, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value  # B列=收支类型
        c = ws.cell(row=r, column=3).value  # C列=金额
        if b is None or c is None:
            continue
        b_str = str(b).strip()
        try:
            amount = float(c)
        except (ValueError, TypeError):
            continue
        if '收入' in b_str:
            balance += amount
        elif '支出' in b_str:
            balance -= amount
    return round(balance, 2)


# ==================== API 路由 ====================

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')


@app.route('/api/sheets')
def list_sheets():
    all_sheets = []
    for s in BZR_SHEETS:
        all_sheets.append({'name': s, 'category': '教师工作台', 'file': '班主任工作台.xlsx'})
    for s in HEALTH_SHEETS:
        all_sheets.append({'name': s, 'category': '个人工作台', 'file': '健康追踪表.xlsx'})
    return jsonify(all_sheets)


@app.route('/api/sheet/<name>')
def get_sheet(name):
    wb, ws = _get_wb(name)
    if wb is None:
        return jsonify({'error': f'工作表 "{name}" 不存在'}), 404
    try:
        data = _read_sheet(ws)
        wb.close()
        return jsonify(data)
    except Exception as e:
        wb.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sheet/<name>/append', methods=['POST'])
def append_row(name):
    wb, ws = _get_wb(name)
    if wb is None:
        return jsonify({'error': f'工作表 "{name}" 不存在'}), 404

    row_data = request.json.get('data', [])
    if not row_data:
        return jsonify({'error': '缺少 data 参数'}), 400

    try:
        data_start = 4
        insert_row = data_start
        for r in range(data_start, ws.max_row + 2):
            empty = True
            for c in range(1, min(ws.max_column + 1, 4)):  # 检查前3列
                if ws.cell(row=r, column=c).value is not None:
                    empty = False
                    break
            if empty:
                insert_row = r
                break

        for i, val in enumerate(row_data):
            if i >= ws.max_column:
                break
            ws.cell(row=insert_row, column=i + 1, value=val)

        wb.save(SHEET_FILE[name])
        wb.close()
        return jsonify({'ok': True, 'row': insert_row})
    except Exception as e:
        wb.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/sheet/<name>/update', methods=['PUT'])
def update_cell(name):
    wb, ws = _get_wb(name)
    if wb is None:
        return jsonify({'error': f'工作表 "{name}" 不存在'}), 404

    row = request.json.get('row')
    col = request.json.get('col')
    value = request.json.get('value')

    if row is None or col is None:
        return jsonify({'error': '缺少 row 或 col 参数'}), 400

    try:
        ws.cell(row=row, column=col, value=value)
        wb.save(SHEET_FILE[name])
        wb.close()
        return jsonify({'ok': True, 'row': row, 'col': col})
    except Exception as e:
        wb.close()
        return jsonify({'error': str(e)}), 500


# ==================== 导出 ====================

@app.route('/api/export/<path:name>')
def export_sheet(name):
    """下载 Excel 原文件"""
    fp = SHEET_FILE.get(name)
    if not fp or not os.path.exists(fp):
        return jsonify({'error': f'工作表 "{name}" 不存在'}), 404
    return send_file(fp, as_attachment=True, download_name=os.path.basename(fp))


# ==================== 知识库 ====================

@app.route('/api/knowledge/notes')
def list_notes():
    """列出知识库所有 Markdown 笔记"""
    notes = []
    if not os.path.exists(KB_DIR):
        return jsonify({'notes': [], 'categories': []})
    for root, dirs, files in os.walk(KB_DIR):
        for f in files:
            if f.endswith('.md'):
                full_path = os.path.join(root, f)
                rel_path = os.path.relpath(full_path, KB_DIR)
                parts = rel_path.replace('\\', '/').split('/')
                category = parts[0] if len(parts) > 1 else '根目录'
                name = os.path.splitext(parts[-1])[0]
                stat = os.stat(full_path)
                notes.append({
                    'name': name,
                    'filename': f,
                    'category': category,
                    'relative_path': rel_path,
                    'size': stat.st_size,
                    'modified': stat.st_mtime
                })
    categories = [d for d in os.listdir(KB_DIR) if os.path.isdir(os.path.join(KB_DIR, d))]
    notes.sort(key=lambda x: x['modified'], reverse=True)
    return jsonify({'notes': notes, 'categories': categories})

@app.route('/api/knowledge/create', methods=['POST'])
def create_note():
    """在知识库中创建新笔记"""
    data = request.json or {}
    category = data.get('category', '个人成长')
    title = data.get('title', '').strip()
    content = data.get('content', '')
    template = data.get('template', '')

    if not title:
        return jsonify({'error': '请输入笔记标题'}), 400

    category_dir = os.path.join(KB_DIR, category)
    os.makedirs(category_dir, exist_ok=True)

    safe_title = title.replace('/', '-').replace('\\', '-')
    filename = f'{safe_title}.md'
    filepath = os.path.join(category_dir, filename)

    if os.path.exists(filepath):
        return jsonify({'error': f'笔记 "{title}" 已存在'}), 409

    today = datetime.now().strftime('%Y-%m-%d')
    full_content = f'---\ntitle: {title}\ndate: {today}\ncategory: {category}\ntags: []\n---\n\n'

    if template == '备课笔记':
        full_content += '# 备课笔记\n\n## 课题\n\n## 教学目标\n\n- 知识目标：\n- 能力目标：\n- 情感目标：\n\n## 教学重难点\n\n**重点：**\n\n**难点：**\n\n## 教学过程\n\n### 导入\n\n### 新课讲授\n\n### 课堂小结\n\n### 作业布置\n\n## 教学反思\n\n'
    elif template == '考研知识点':
        full_content += '# 考研知识点\n\n## 所属科目\n\n## 知识点概述\n\n## 核心概念\n\n1. \n2. \n3. \n\n## 记忆口诀\n\n## 真题链接\n\n- [ ] 年份/题型：\n\n## 复习记录\n\n| 日期 | 掌握程度 | 备注 |\n|------|----------|------|\n| {today} | 初次学习 | |\n'
    elif template == '读书笔记':
        full_content += '# 读书笔记\n\n## 书籍信息\n\n- 书名：\n- 作者：\n- 阅读日期：{today}\n\n## 核心观点\n\n## 精彩摘录\n\n> \n\n## 我的思考\n\n## 行动清单\n\n- [ ] \n'
    elif template == '学生档案':
        full_content += '# 学生档案\n\n## 基本信息\n\n- 姓名：\n- 学号：\n- 家庭情况：\n\n## 学业表现\n\n## 行为记录\n\n| 日期 | 事件 | 处理 |\n|------|------|------|\n| {today} | | |\n\n## 重点关注\n\n'
    elif template == '班会记录':
        full_content += '# 班会记录\n\n- 日期：{today}\n- 主题：\n- 主持人：\n\n## 会议内容\n\n## 学生反馈\n\n## 后续跟进\n\n'
    elif template == '班主任日志':
        full_content += '# 班主任日志\n\n- 日期：{today}\n- 天气：\n\n## 今日记事\n\n## 好人好事\n\n## 存在问题\n\n## 明日计划\n\n'

    full_content += content

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(full_content)

    return jsonify({'ok': True, 'path': os.path.relpath(filepath, KB_DIR), 'title': title})

@app.route('/api/knowledge/open/<path:relpath>')
def open_in_obsidian(relpath):
    """返回 obsidian:// 协议链接，点击在 Obsidian 中打开"""
    # Obsidian vault name 使用知识库文件夹名
    vault_name = '知识库'
    note_name = relpath.replace('.md', '').replace('\\', '/')
    uri = f'obsidian://open?vault={vault_name}&file={note_name}'
    return jsonify({'uri': uri})


# ==================== 统计接口 ====================

@app.route('/api/stats/dashboard')
def dashboard_stats():
    stats = {
        'total_students': 0,
        'today_attendance': {'出勤': 0, '迟到': 0, '请假': 0, '缺勤': 0},
        'top_points': [],
        'recent_logs': [],
        'class_fund_balance': 0
    }

    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['学生信息总表']
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=1).value is not None:
                stats['total_students'] += 1

        # 积分排行 —— 手动计算月合计（C-J列求和）
        ws_points = wb['日常行为积分']
        students = []
        for r in range(4, ws_points.max_row + 1):
            name = ws_points.cell(row=r, column=2).value
            if not name:
                continue
            total = _sum_range(ws_points, r, 3, 10) or 0  # C-J列(第3-10列)
            students.append({'name': str(name), 'points': int(total)})
        students.sort(key=lambda x: x['points'], reverse=True)
        stats['top_points'] = students[:5]

        # 考勤统计
        ws_att = wb['考勤管理']
        for r in range(4, ws_att.max_row + 1):
            status = ws_att.cell(row=r, column=5).value
            if status:
                s = str(status).strip()
                if s in stats['today_attendance']:
                    stats['today_attendance'][s] += 1

        # 班费余额 —— 手动计算，不读G列公式
        ws_fund = wb['班费管理']
        stats['class_fund_balance'] = _calc_balance(ws_fund)

        # 最近日志
        ws_log = wb['班主任日志']
        logs = []
        for r in range(4, ws_log.max_row + 1):
            date_val = ws_log.cell(row=r, column=1).value
            content = ws_log.cell(row=r, column=4).value
            if date_val and content:
                logs.append({'date': str(date_val), 'content': str(content)[:50]})
        stats['recent_logs'] = logs[-5:]

        wb.close()
    except Exception as e:
        stats['error'] = str(e)

    return jsonify(stats)


@app.route('/api/stats/attendance')
def attendance_stats():
    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['考勤管理']
        status_count = {}
        for r in range(4, ws.max_row + 1):
            status = ws.cell(row=r, column=5).value
            if status:
                s = str(status).strip()
                status_count[s] = status_count.get(s, 0) + 1

        date_count = {}
        for r in range(4, ws.max_row + 1):
            date_val = ws.cell(row=r, column=3).value
            status = ws.cell(row=r, column=5).value
            if date_val and status:
                d = str(date_val)[:10]
                if d not in date_count:
                    date_count[d] = {'出勤': 0, '迟到': 0, '请假': 0, '缺勤': 0, '总人数': 0}
                s = str(status).strip()
                if s in date_count[d]:
                    date_count[d][s] += 1
                date_count[d]['总人数'] += 1

        wb.close()
        return jsonify({'status_count': status_count, 'date_stats': date_count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/scores')
def scores_stats():
    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['成绩跟踪']

        students = []
        for r in range(4, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if not name:
                continue

            # 月考1各科 (C-H, 第3-8列) —— 手动求和
            yuekao_scores = []
            for c in range(3, 9):
                yuekao_scores.append(_safe_number(ws, r, c))
            yk_total = _sum_range(ws, r, 3, 8)

            # 期中各科 (K-P, 第11-16列) —— 手动求和
            qizhong_scores = []
            for c in range(11, 17):
                qizhong_scores.append(_safe_number(ws, r, c))
            qz_total = _sum_range(ws, r, 11, 16)

            # 排名和进退步
            rank_val = _safe_number(ws, r, 18)
            change_val = _safe_number(ws, r, 19)

            students.append({
                'name': str(name),
                'yuekao1': yuekao_scores,
                'yuekao1_total': yk_total,
                'qizhong': qizhong_scores,
                'qizhong_total': qz_total,
                'rank': rank_val,
                'change': change_val,
            })

        subjects = ['语文', '数学', '英语', '政治', '历史', '地理']
        avg_scores = {'yuekao1': {}, 'qizhong': {}}
        for i, subj in enumerate(subjects):
            yk_vals = [s['yuekao1'][i] for s in students if s['yuekao1'][i] is not None]
            qz_vals = [s['qizhong'][i] for s in students if s['qizhong'][i] is not None]
            avg_scores['yuekao1'][subj] = round(sum(yk_vals) / len(yk_vals), 1) if yk_vals else 0
            avg_scores['qizhong'][subj] = round(sum(qz_vals) / len(qz_vals), 1) if qz_vals else 0

        wb.close()
        return jsonify({'students': students, 'avg_scores': avg_scores, 'subjects': subjects})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/points')
def points_stats():
    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['日常行为积分']

        students = []
        for r in range(4, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if not name:
                continue
            weekly = []
            for c in range(3, 11):  # C-J: 8周
                val = _safe_number(ws, r, c)
                weekly.append(val if val is not None else 0)
            total = _sum_range(ws, r, 3, 10) or sum(weekly)

            students.append({'name': str(name), 'weekly': weekly, 'total': int(total)})

        students.sort(key=lambda x: x['total'], reverse=True)
        wb.close()
        return jsonify({'students': students})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/seating')
def seating_data():
    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['座位表']

        grid = []
        special_cells = {}
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is not None:
                    val_str = str(val)
                    if val_str in ['讲台', '前门', '后门', '过道']:
                        special_cells[f'{r},{c}'] = val_str
                    row_data.append(val_str)
                else:
                    row_data.append('')
            grid.append(row_data)

        wb.close()
        return jsonify({'grid': grid, 'specials': special_cells, 'rows': len(grid), 'cols': len(grid[0]) if grid else 0})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/seating/update', methods=['POST'])
def seating_update():
    data = request.json
    row = data.get('row')
    col = data.get('col')
    value = data.get('value')

    try:
        wb = load_workbook(XLSX_BANZHUREN, data_only=False)
        ws = wb['座位表']
        ws.cell(row=row, column=col, value=value)
        wb.save(XLSX_BANZHUREN)
        wb.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print('美美大王工作台 v2.1 启动中...')
    print('浏览器打开 http://localhost:5000')
    app.run(host='127.0.0.1', port=5000, debug=True)
