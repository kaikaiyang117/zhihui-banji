# -*- coding: utf-8 -*-
"""一次性数据迁移：旧 Excel（班主任工作台.xlsx / 健康追踪表.xlsx）→ SQLite

公式列（总分/排名/余额/腰臀比等）不再入库，由读取时动态计算。
"""
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from app import db
from app.config import (LEGACY_BZR, LEGACY_HEALTH, SHEET_META, STUDENT_COLUMNS)
from app.services import attendance, funds, points
from app.services.class_context import enroll_student, scope_ids


def _formula(val):
    return isinstance(val, str) and val.startswith('=')


def _norm_header(val):
    return str(val).replace('\n', '').strip() if val is not None else ''


def _clean(v):
    """公式列置空（派生列迁移后动态计算）；原样返回其他值"""
    if _formula(v):
        return None
    return v


def migrate_file(fp: str, sheet_names: list[str]):
    wb = load_workbook(fp, data_only=False)
    for name in sheet_names:
        if name not in wb.sheetnames:
            print(f'  跳过（不存在）: {name}')
            continue
        ws = wb[name]
        headers = [_norm_header(ws.cell(row=3, column=c).value)
                   for c in range(1, ws.max_column + 1)]
        while headers and headers[-1] == '':
            headers.pop()
        category = SHEET_META.get(name, {}).get('category', '')
        group = SHEET_META.get(name, {}).get('group', 'teacher')
        db.set_sheet_meta(name, headers, category, group)
        n = 0
        if name == '座位表':
            class_id, term_id = scope_ids(write=True, conn=db.get_conn())
            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        db.get_conn().execute(
                            'INSERT OR REPLACE INTO seating(class_id,term_id,r,c,val) VALUES(?,?,?,?,?)',
                            (class_id, term_id, r, c, str(v)))
            db.get_conn().commit()
            print(f'  座位表: 迁移完成')
            continue
        for r in range(4, ws.max_row + 1):
            row = []
            has_data = False
            for c in range(1, len(headers) + 1):
                v = _clean(ws.cell(row=r, column=c).value)
                row.append(v)
                if v is not None:
                    has_data = True
            if has_data:
                if name == '考勤管理':
                    index = {header: i for i, header in enumerate(headers)}
                    xh = str(row[index.get('学号', -1)] or '').strip() if '学号' in index else ''
                    student_name = str(row[index.get('姓名', -1)] or '').strip() if '姓名' in index else ''
                    class_id, term_id = scope_ids(conn=db.get_conn())
                    student = db.get_conn().execute(
                        '''SELECT s.id FROM students s JOIN student_enrollments e ON e.student_id=s.id
                           WHERE e.class_id=? AND e.term_id=?
                             AND ((?<>'' AND s.学号=?) OR (?='' AND s.姓名=?))
                           ORDER BY s.id LIMIT 1''',
                        (class_id, term_id, xh, xh, xh, student_name),
                    ).fetchone()
                    if not student:
                        continue
                    value = lambda key: row[index[key]] if key in index and index[key] < len(row) else ''
                    attendance.save_daily(
                        str(value('日期') or '')[:10], '常规到校', [{
                            'student_id': student['id'], 'status': str(value('状态') or '出勤'),
                            'arrive': str(value('到校时间') or ''),
                            'leave': str(value('离校时间') or ''),
                            'reason': str(value('原因') or value('请假原因') or ''),
                            'note': str(value('备注') or ''),
                        }], evaluate=False,
                    )
                else:
                    db.insert_row(name, row)
                n += 1
        print(f'  {name}: {n} 行')
    wb.close()


def migrate_students():
    wb = load_workbook(LEGACY_BZR, data_only=False)
    ws = wb['学生信息总表']
    # 旧表头含空格/换行（如「监护人 姓名」「是否 住校」），归一化后映射
    col_by_norm = {}
    for c in range(1, ws.max_column + 1):
        h = _norm_header(ws.cell(row=3, column=c).value).replace(' ', '')
        if h:
            col_by_norm[h] = c
    conn = db.get_conn()
    n = 0
    for r in range(4, ws.max_row + 1):
        xh = ws.cell(row=r, column=col_by_norm['学号']).value if '学号' in col_by_norm else None
        name = ws.cell(row=r, column=col_by_norm['姓名']).value if '姓名' in col_by_norm else None
        if xh is None and name is None:
            continue
        fields = {}
        for col in STUDENT_COLUMNS:
            norm = col.replace(' ', '')
            c = col_by_norm.get(norm)
            fields[col] = _clean(ws.cell(row=r, column=c).value) if c else None
        if not fields['学号']:
            fields['学号'] = str(xh).strip() if xh is not None else ''
        if not fields['姓名'] and name is not None:
            fields['姓名'] = _clean(name)
        if not fields['学号'] or not fields['姓名']:
            continue
        vals = tuple(fields.get(k) for k in STUDENT_COLUMNS)
        cols_sql = ','.join(f'[{k}]' for k in STUDENT_COLUMNS)
        ph = ','.join('?' for _ in STUDENT_COLUMNS)
        student_id = conn.execute(f'INSERT INTO students({cols_sql}) VALUES({ph})', vals).lastrowid
        enroll_student(student_id, conn=conn, commit=False)
        n += 1
    conn.commit()
    wb.close()
    print(f'学生信息总表: {n} 名学生')


def main():
    os.makedirs(os.path.dirname(db.DB_PATH), exist_ok=True)
    if os.path.exists(db.DB_PATH):
        os.replace(db.DB_PATH, db.DB_PATH + '.bak')
    db.get_conn()
    print('=== 迁移学生信息 ===')
    migrate_students()
    print('=== 迁移班主任工作台 ===')
    bzr = [n for n, m in SHEET_META.items() if m['group'] == 'teacher']
    migrate_file(LEGACY_BZR, bzr)
    print('=== 迁移行为积分流水 ===')
    report = points.migrate_legacy_rows(conn=db.get_conn())
    print(f"  行为积分: 导入 {report['imported_entries']} 条，跳过 {report['skipped_entries']} 条")
    print('=== 迁移班费分类账 ===')
    fund_report = funds.migrate_legacy_rows(conn=db.get_conn())
    print(f"  班费管理: 导入 {fund_report['imported_entries']} 条，跳过 {fund_report['skipped_entries']} 条")
    print('=== 迁移健康追踪表 ===')
    health = [n for n, m in SHEET_META.items() if m['group'] == 'personal']
    migrate_file(LEGACY_HEALTH, health)
    print('=== 迁移完成 ===')


if __name__ == '__main__':
    main()
