# -*- coding: utf-8 -*-
"""P1-A/B/C：搜索、成绩、考勤规则、班级任务与值日安排。"""
from __future__ import annotations

import io
from typing import Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from .. import db
from ..services import (
    attendance as attendance_service,
    class_context,
    class_tasks as class_tasks_service,
    duty as duty_service,
    scores as scores_service,
    work_items,
)

router = APIRouter(prefix='/api')


def _rows(sql: str, params: tuple = ()) -> list[dict]:
    return [dict(row) for row in db.get_conn().execute(sql, params).fetchall()]


def _scope(write: bool = False) -> tuple[int, int]:
    return class_context.scope_ids(write=write, conn=db.get_conn())


def _student(student_id: int, write: bool = False) -> dict:
    try:
        return class_context.ensure_student_in_scope(student_id, write=write, conn=db.get_conn())
    except class_context.ArchivedScopeError:
        raise
    except class_context.ScopeError as exc:
        raise HTTPException(404, str(exc)) from exc


class ExamRecord(BaseModel):
    student_id: int
    exam_name: str = Field(min_length=1)
    exam_date: str = ''
    subject: str = Field(min_length=1)
    score: Optional[float] = None
    rank: Optional[int] = None


def upsert_exam_record(record: ExamRecord) -> int:
    scores_service.commit_exam_rows([{
        'row': 1, 'valid': True, 'student_id': record.student_id,
        'exam_name': record.exam_name, 'exam_date': record.exam_date,
        'subject': record.subject, 'score': record.score, 'rank': record.rank,
        'record_status': '正常', 'note': '',
    }], filename='单条成绩', request_id=f'single-{record.student_id}-{record.exam_name}-{record.subject}-{record.score}-{record.rank}')
    class_id, term_id = _scope()
    return int(db.get_conn().execute(
        '''SELECT id FROM exam_records WHERE student_id=? AND class_id=? AND term_id=?
           AND exam_name=? AND subject=? AND deleted_at='' ''',
        (record.student_id, class_id, term_id, record.exam_name, record.subject),
    ).fetchone()['id'])


def import_exam_rows(rows: list[list]) -> dict:
    return scores_service.import_exam_rows(rows)


@router.post('/exams/import')
async def import_exams(file: UploadFile = File(...)):
    """兼容入口现在只做预览，不再收到文件后立即写库。"""
    return await preview_exams(file, 'update')


@router.post('/exams/import/preview')
async def preview_exams(file: UploadFile = File(...), duplicate_strategy: str = 'update'):
    try:
        workbook = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
        sheet = workbook.active
        result = scores_service.preview_exam_rows(
            [list(row) for row in sheet.iter_rows(values_only=True)], duplicate_strategy)
        workbook.close()
        return {'ok': True, 'filename': file.filename or '', **result}
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc
    except Exception as exc:
        raise HTTPException(400, f'成绩文件读取失败：{exc}') from exc


class ExamImportCommitBody(BaseModel):
    filename: str = ''
    duplicate_strategy: str = 'update'
    request_id: str = ''
    rows: list[dict] = []


@router.post('/exams/import/commit')
def commit_exams(body: ExamImportCommitBody):
    try:
        return scores_service.commit_exam_rows(
            body.rows, filename=body.filename,
            duplicate_strategy=body.duplicate_strategy, request_id=body.request_id)
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.get('/exams')
def list_exams(student_id: Optional[int] = None, exam_name: Optional[str] = None):
    records = scores_service.list_records(student_id=student_id)
    if exam_name:
        records = [item for item in records if item['exam_name'] == exam_name]
    return {'records': records}


@router.get('/exams/summary')
def exam_summary(student_id: Optional[int] = None):
    return scores_service.score_summary(student_id=student_id)


class ScoreSubjectBody(BaseModel):
    name: str = Field(min_length=1)
    full_score: float = Field(default=0, ge=0)
    enabled: bool = True
    sort_order: int = 0
    subject_group: str = '必考'
    score_type: str = '原始分'


class ScoreSubjectUpdate(BaseModel):
    name: Optional[str] = None
    full_score: Optional[float] = Field(default=None, ge=0)
    enabled: Optional[bool] = None
    sort_order: Optional[int] = None
    subject_group: Optional[str] = None
    score_type: Optional[str] = None


class ScoreExamBody(BaseModel):
    name: str = Field(min_length=1)
    exam_date: str = ''
    subject_ids: list[int] = []
    enabled: bool = True
    sort_order: int = 0


class ScoreExamUpdate(BaseModel):
    name: Optional[str] = None
    exam_date: Optional[str] = None
    subject_ids: Optional[list[int]] = None
    enabled: Optional[bool] = None
    sort_order: Optional[int] = None


@router.get('/score-config')
def score_config():
    return scores_service.list_config()


@router.post('/score-config/subjects')
def add_score_subject(body: ScoreSubjectBody):
    try:
        return scores_service.create_subject(**body.model_dump())
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/score-config/subjects/{subject_id}')
def edit_score_subject(subject_id: int, body: ScoreSubjectUpdate):
    try:
        return scores_service.update_subject(
            subject_id, **body.model_dump(exclude_none=True))
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


class ScoreSettingsBody(BaseModel):
    mode: str = '固定科目'


class StudentScoreSubjectsBody(BaseModel):
    subject_ids: list[int] = []


@router.put('/score-config/settings')
def edit_score_settings(body: ScoreSettingsBody):
    try:
        return scores_service.update_term_settings(mode=body.mode)
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/score-config/students/{student_id}/subjects')
def edit_student_score_subjects(student_id: int, body: StudentScoreSubjectsBody):
    try:
        return scores_service.save_student_subjects(student_id, body.subject_ids)
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post('/score-config/exams')
def add_score_exam(body: ScoreExamBody):
    try:
        return scores_service.create_exam(**body.model_dump())
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/score-config/exams/{exam_id}')
def edit_score_exam(exam_id: int, body: ScoreExamUpdate):
    try:
        return scores_service.update_exam(exam_id, **body.model_dump(exclude_none=True))
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


class ScoreRuleBody(BaseModel):
    name: str = Field(min_length=1)
    metric: str = '总分下降'
    subject_id: Optional[int] = None
    threshold: float = Field(default=10, gt=0)
    priority: str = '重要'
    enabled: bool = True


class ScoreRuleUpdate(BaseModel):
    enabled: Optional[bool] = None
    threshold: Optional[float] = Field(default=None, gt=0)
    priority: Optional[str] = None


@router.get('/score-rules')
def get_score_rules(source_id: Optional[int] = None):
    return scores_service.list_rules(source_id=source_id)


@router.post('/score-rules')
def add_score_rule(body: ScoreRuleBody):
    try:
        return scores_service.create_rule(**body.model_dump())
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/score-rules/{rule_id}')
def edit_score_rule(rule_id: int, body: ScoreRuleUpdate):
    try:
        return scores_service.update_rule(rule_id, **body.model_dump(exclude_none=True))
    except scores_service.ScoreError as exc:
        status = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status, str(exc)) from exc


@router.post('/score-rules/evaluate')
def evaluate_score_rules():
    try:
        return scores_service.evaluate_rules(trigger='manual')
    except scores_service.ScoreError as exc:
        raise HTTPException(400, str(exc)) from exc


class AttendanceRuleBody(BaseModel):
    name: str = Field(min_length=1)
    metric: str = '迟到次数'
    threshold: int = Field(default=2, ge=1)
    period_days: int = Field(default=7, ge=1, le=365)
    priority: str = '重要'
    enabled: bool = True
    scene: str = '全部场景'


class AttendanceRuleUpdate(BaseModel):
    enabled: Optional[bool] = None
    threshold: Optional[int] = Field(default=None, ge=1)
    period_days: Optional[int] = Field(default=None, ge=1, le=365)
    priority: Optional[str] = None
    scene: Optional[str] = None


class AttendanceRuleEvaluation(BaseModel):
    reference_date: str = ''


@router.get('/attendance/rules')
def list_attendance_rules(source_id: Optional[int] = None):
    return attendance_service.list_rules(source_id=source_id)


@router.post('/attendance/rules')
def create_attendance_rule(body: AttendanceRuleBody):
    try:
        return attendance_service.create_rule(**body.model_dump())
    except attendance_service.AttendanceError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/attendance/rules/{rule_id}')
def update_attendance_rule(rule_id: int, body: AttendanceRuleUpdate):
    try:
        return attendance_service.update_rule(
            rule_id, **body.model_dump(exclude_none=True))
    except attendance_service.AttendanceError as exc:
        status = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status, str(exc)) from exc


@router.post('/attendance/rules/evaluate')
def evaluate_attendance_rules(body: AttendanceRuleEvaluation | None = None):
    try:
        result = attendance_service.evaluate_rules(
            reference_date=body.reference_date if body else '', trigger='manual')
        return {**result, 'created': [
            item for item in result['summary'] if item['state'] == '新命中'
        ], 'count': result['created_count']}
    except attendance_service.AttendanceError as exc:
        raise HTTPException(400, str(exc)) from exc


class ClassTaskBody(BaseModel):
    title: str = Field(min_length=1)
    task_type: str = '材料收集'
    start_at: str = ''
    due_at: str = ''
    material_name: str = ''
    description: str = ''
    student_ids: list[int] = Field(default_factory=list)
    template_id: Optional[int] = None


class ClassTaskUpdate(BaseModel):
    status: Optional[str] = None
    description: Optional[str] = None
    start_at: Optional[str] = None
    due_at: Optional[str] = None
    completion_result: Optional[str] = None
    confirm_incomplete: bool = False


class ClassTaskItemUpdate(BaseModel):
    status: str = '已提交'
    note: str = ''


class ClassTaskReminderBody(BaseModel):
    student_ids: Optional[list[int]] = None


class ClassTaskTemplateBody(BaseModel):
    name: str = Field(min_length=1)
    task_type: str = '材料收集'
    material_name: str = ''
    description: str = ''
    default_due_days: int = Field(default=7, ge=0, le=366)
    enabled: bool = True


class ClassTaskTemplateUpdate(BaseModel):
    name: Optional[str] = None
    task_type: Optional[str] = None
    material_name: Optional[str] = None
    description: Optional[str] = None
    default_due_days: Optional[int] = Field(default=None, ge=0, le=366)
    enabled: Optional[bool] = None


@router.get('/class-tasks')
def list_class_tasks(status: Optional[str] = None, timing_state: str = '', source_id: Optional[int] = None):
    try:
        return {'tasks': class_tasks_service.list_tasks(
            status=status or '', timing_state=timing_state, source_id=source_id)}
    except class_tasks_service.ClassTaskError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.get('/class-task-templates')
def list_class_task_templates(include_disabled: bool = False):
    try:
        return {'templates': class_tasks_service.list_templates(include_disabled=include_disabled)}
    except class_tasks_service.ClassTaskError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post('/class-task-templates')
def create_class_task_template(body: ClassTaskTemplateBody):
    try:
        return {'ok': True, 'template': class_tasks_service.create_template(**body.model_dump())}
    except class_tasks_service.ClassTaskError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.put('/class-task-templates/{template_id}')
def update_class_task_template(template_id: int, body: ClassTaskTemplateUpdate):
    try:
        return {'ok': True, 'template': class_tasks_service.update_template(
            template_id, **body.model_dump(exclude_none=True))}
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.post('/class-tasks')
def create_class_task(body: ClassTaskBody):
    try:
        task = class_tasks_service.create_task(**body.model_dump())
        return {'ok': True, 'task_id': task['id'], 'task': task}
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '学生' in str(exc) and '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.put('/class-tasks/{task_id}')
def update_class_task(task_id: int, body: ClassTaskUpdate):
    try:
        return {'ok': True, 'task': class_tasks_service.update_task(
            task_id, **body.model_dump(exclude_none=True))}
    except class_tasks_service.IncompleteTaskError as exc:
        raise HTTPException(409, {'message': str(exc), 'missing_students': exc.missing}) from exc
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.put('/class-tasks/{task_id}/items/{student_id}')
def update_class_task_item(task_id: int, student_id: int, body: ClassTaskItemUpdate):
    try:
        return {'ok': True, 'task': class_tasks_service.update_item(
            task_id, student_id, status=body.status, note=body.note)}
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '没有该学生' in str(exc) or '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.post('/class-tasks/{task_id}/remind')
def remind_class_task(task_id: int, body: ClassTaskReminderBody | None = None):
    try:
        return class_tasks_service.remind(task_id, body.student_ids if body else None)
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.post('/class-tasks/{task_id}/attachments/{student_id}')
async def upload_class_task_attachment(task_id: int, student_id: int, file: UploadFile = File(...)):
    try:
        content = await file.read()
        attachment = class_tasks_service.save_attachment(
            task_id, student_id, original_name=file.filename or '附件',
            content_type=file.content_type or 'application/octet-stream', content=content)
        return {'ok': True, 'attachment': attachment}
    except class_tasks_service.ClassTaskError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.get('/class-tasks/attachments/{attachment_id}')
def download_class_task_attachment(attachment_id: int):
    try:
        attachment, path = class_tasks_service.attachment_file(attachment_id)
    except class_tasks_service.ClassTaskError as exc:
        raise HTTPException(404, str(exc)) from exc
    return FileResponse(path, media_type=attachment['content_type'], filename=attachment['original_name'])


class DutyBody(BaseModel):
    duty_date: str = Field(min_length=10)
    area: str = Field(min_length=1)
    student_id: int
    status: str = '待完成'
    note: str = ''
    completion_result: str = ''


@router.get('/duty')
def list_duty(duty_date: Optional[str] = None, date_from: str = '', date_to: str = '', source_id: Optional[int] = None):
    try:
        return {'assignments': duty_service.list_assignments(
            duty_date=duty_date or '', date_from=date_from, date_to=date_to, source_id=source_id)}
    except duty_service.DutyError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post('/duty')
def create_duty(body: DutyBody):
    try:
        assignment = duty_service.create_assignment(**body.model_dump())
        return {'ok': True, 'assignment_id': assignment['id'], 'assignment': assignment}
    except duty_service.DutyConflictError as exc:
        raise HTTPException(409, {'message': str(exc), 'conflicts': exc.conflicts}) from exc
    except duty_service.DutyError as exc:
        status_code = 404 if '学生' in str(exc) and '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


class DutyUpdate(BaseModel):
    status: str
    note: str = ''
    completion_result: str = ''


class DutyRotationRuleBody(BaseModel):
    name: str = Field(min_length=1)
    area: str = Field(min_length=1)
    start_date: str = Field(min_length=10)
    end_date: str = ''
    weekday_mask: int = Field(default=31, ge=1, le=127)
    student_ids: list[int] = Field(default_factory=list)
    enabled: bool = True


class DutyGenerateBody(BaseModel):
    date_from: str = ''
    date_to: str = ''
    confirm: bool = False


@router.put('/duty/{assignment_id}')
def update_duty(assignment_id: int, body: DutyUpdate):
    try:
        return {'ok': True, 'assignment': duty_service.update_assignment(
            assignment_id, status=body.status, note=body.note,
            completion_result=body.completion_result)}
    except duty_service.DutyError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.get('/duty/rotation-rules')
def list_duty_rotation_rules(include_disabled: bool = False):
    try:
        return {'rules': duty_service.list_rotation_rules(include_disabled=include_disabled)}
    except duty_service.DutyError as exc:
        raise HTTPException(400, str(exc)) from exc


@router.post('/duty/rotation-rules')
def create_duty_rotation_rule(body: DutyRotationRuleBody):
    try:
        return {'ok': True, 'rule': duty_service.create_rotation_rule(**body.model_dump())}
    except duty_service.DutyError as exc:
        status_code = 404 if '学生' in str(exc) and '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.post('/duty/rotation-rules/{rule_id}/generate')
def generate_duty_rotation(rule_id: int, body: DutyGenerateBody | None = None):
    body = body or DutyGenerateBody()
    try:
        return duty_service.generate_rotation(
            rule_id, date_from=body.date_from, date_to=body.date_to, confirm=body.confirm)
    except duty_service.DutyConflictError as exc:
        raise HTTPException(409, {'message': str(exc), 'conflicts': exc.conflicts}) from exc
    except duty_service.DutyError as exc:
        status_code = 404 if '不存在' in str(exc) else 400
        raise HTTPException(status_code, str(exc)) from exc


@router.get('/search')
def search(q: str = '', limit: int = 30):
    query = q.strip()
    if not query:
        return {'results': []}
    like = f'%{query}%'
    class_id, term_id = _scope()
    results = []
    for row in _rows(
        'SELECT s.id, s.学号, s.姓名, s.班级任职 FROM students s '
        'JOIN student_enrollments e ON e.student_id=s.id '
        "WHERE e.class_id=? AND e.term_id=? AND e.status='在读' AND s.deleted_at='' "
        'AND (s.学号 LIKE ? OR s.姓名 LIKE ? OR s.备注 LIKE ?) '
        'ORDER BY s.学号 LIMIT ?', (class_id, term_id, like, like, like, limit)):
        results.append({'kind': 'student', 'id': row['id'], 'student_id': row['id'], 'title': row['姓名'], 'summary': f"{row['学号'] or '暂无学号'} · {row['班级任职'] or '班级成员'}", 'path': f"/student/{row['id']}"})
    sources = [
        ('student_events', 'x.id, x.student_id, x.event_type AS title, x.description AS summary', '事件', '/events'),
        ('student_tasks', 'x.id, x.student_id, x.title, x.notes AS summary', '待办', '/tasks'),
        ('focus_items', 'x.id, x.student_id, x.topic AS title, x.reason AS summary', '关注', '/special'),
        ('communications', 'x.id, x.student_id, x.reason AS title, x.summary', '沟通', '/parent-comm'),
    ]
    for table, fields, kind, path in sources:
        for row in _rows(
            f'SELECT {fields}, s.姓名 AS student_name FROM {table} x JOIN students s ON s.id=x.student_id '
            "WHERE x.class_id=? AND x.term_id=? AND x.deleted_at='' AND s.deleted_at='' AND (title LIKE ? OR summary LIKE ?) LIMIT ?",
            (class_id, term_id, like, like, limit)):
            results.append({'kind': kind, 'id': row['id'], 'student_id': row['student_id'], 'title': row['title'], 'summary': f"{row['student_name']} · {row['summary'] or ''}", 'path': path})
    for row in _rows(
        'SELECT e.id, e.student_id, e.exam_name AS title, e.subject || " " || COALESCE(e.score, "") AS summary, '
        's.姓名 AS student_name FROM exam_records e JOIN students s ON s.id=e.student_id '
        "WHERE e.class_id=? AND e.term_id=? AND e.deleted_at='' AND s.deleted_at='' AND (e.exam_name LIKE ? OR e.subject LIKE ?) LIMIT ?",
        (class_id, term_id, like, like, limit)):
        results.append({'kind': '成绩', 'id': row['id'], 'student_id': row['student_id'], 'title': row['title'], 'summary': f"{row['student_name']} · {row['summary']}", 'path': '/scores'})
    return {'results': results[:limit]}
