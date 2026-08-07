# -*- coding: utf-8 -*-
"""Excel 导出：通用工作表导出 + 汇总报表"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import STUDENT_COLUMNS
from .db import get_conn, get_rows, get_sheet_meta
from .derived import derive
from .services import attendance, points, scores
from .services.class_context import scope_ids

HEADER_FILL = PatternFill('solid', fgColor='5B6ABF')
HEADER_FONT = Font(color='FFFFFF', bold=True)


def _sheet_bytes(title: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_sheet(sheet: str) -> tuple[io.BytesIO, str]:
    """通用工作表导出（含派生计算列）"""
    if sheet == '座位表':
        return export_seating()
    if sheet == '考勤管理':
        rows = attendance.compatibility_rows()
        headers = ['日期', '星期', '学号', '姓名', '状态', '到校时间', '离校时间',
                   '原因', '备注', '考勤场景']
        return _sheet_bytes(sheet, headers, [row['data'] for row in rows]), f'{sheet}.xlsx'
    if sheet == '日常行为积分':
        return export_points()
    meta = get_sheet_meta(sheet)
    headers = meta['headers'] if meta else []
    rows = derive(sheet, get_rows(sheet))
    return _sheet_bytes(sheet, headers, [r['data'] for r in rows]), f'{sheet}.xlsx'


def export_points() -> tuple[io.BytesIO, str]:
    """导出完整积分流水，保留已撤销记录以便核对；排名只使用有效流水。"""
    entries = points.list_entries(limit=5_000)
    headers = ['日期', '周期', '学号', '姓名', '分类', '分值', '原因', '状态',
               '撤销原因', '来源', '规则']
    rows = [[
        item.get('occurred_at') or '历史快照', item.get('period_key', ''),
        item.get('学号', ''), item.get('student_name', ''), item.get('category', ''),
        item.get('amount', 0), item.get('reason', ''), item.get('status', ''),
        item.get('reversal_reason', ''), item.get('source_label', ''), item.get('rule_name', ''),
    ] for item in entries]
    return _sheet_bytes('行为积分流水', headers, rows), '行为积分流水.xlsx'


def export_seating() -> tuple[io.BytesIO, str]:
    """座位表导出（从 seating 表重建网格）"""
    class_id, term_id = scope_ids(conn=get_conn())
    rows = get_conn().execute(
        'SELECT r, c, val FROM seating WHERE class_id=? AND term_id=? ORDER BY r, c',
        (class_id, term_id),
    ).fetchall()
    max_r = max((row['r'] for row in rows), default=0)
    max_c = max((row['c'] for row in rows), default=0)
    grid = [['' for _ in range(max_c + 1)] for _ in range(max_r + 1)]
    for row in rows:
        grid[row['r']][row['c']] = row['val']
    while grid and all(v == '' for v in grid[0]):
        grid.pop(0)
    while grid and all(v == '' for v in grid[-1]):
        grid.pop()
    cols = len(grid[0]) if grid else 0
    for row in grid:
        while len(row) < cols:
            row.append('')
    headers = [f'第{c+1}列' for c in range(cols)]
    return _sheet_bytes('座位表', headers, grid), '座位表.xlsx'


def export_students() -> tuple[io.BytesIO, str]:
    class_id, term_id = scope_ids(conn=get_conn())
    rows = get_conn().execute(
        f'SELECT {",".join("s.["+k+"]" for k in STUDENT_COLUMNS)} FROM students s '
        'JOIN student_enrollments e ON e.student_id=s.id '
        "WHERE e.class_id=? AND e.term_id=? AND e.status='在读' AND s.deleted_at='' ORDER BY s.学号",
        (class_id, term_id),
    ).fetchall()
    data = [[row[k] for k in STUDENT_COLUMNS] for row in rows]
    return _sheet_bytes('学生信息', STUDENT_COLUMNS, data), '学生信息总表.xlsx'


# ---------- 汇总报表 ----------

def export_score_report(exam: str) -> tuple[io.BytesIO, str]:
    """从结构化统计导出任意考试，缺失成绩保持为空而不是按零分处理。"""
    summary = scores.score_summary()
    selected = None
    for item in summary['exams']:
        if str(item['id']) == str(exam) or item['name'] == exam:
            selected = item
            break
    if not selected:
        raise ValueError('考试不存在')
    subject_names = [item['subject'] for item in selected['subject_stats']]
    headers = ['学号', '姓名'] + subject_names + ['总分', '班排名', '分层', '完整性']
    data = []
    for student in summary['students']:
        result = next(
            (item for item in student['exams'] if item['exam_id'] == selected['id']), None)
        if not result or not result['has_any']:
            continue
        subject_values = []
        for name in subject_names:
            item = result['subjects'].get(name)
            if not item:
                subject_values.append(None)
            elif item['status'] == '正常':
                subject_values.append(item['score'])
            else:
                subject_values.append(item['status'])
        data.append([
            student['学号'], student['姓名'], *subject_values, result['total'],
            result['rank'], result['stratum'],
            '完整' if result['complete'] else f"缺少：{'、'.join(result['missing_subjects'])}",
        ])
    name = selected['name']
    return _sheet_bytes(f'成绩汇总-{name}', headers, data), f'成绩汇总_{name}.xlsx'


def export_attendance_report(date_from: str | None, date_to: str | None) -> tuple[io.BytesIO, str]:
    """考勤汇总报表：按日期和场景统计各状态数量。"""
    rows = attendance.list_records(
        date_from=date_from or '', date_to=date_to or '', limit=50_000)
    headers = ['日期', '场景', '出勤', '迟到', '请假', '早退', '缺勤', '总记录']
    grouped: dict[tuple[str, str], dict] = {}
    for row in rows:
        key = (row['attendance_date'], row['scene'])
        bucket = grouped.setdefault(
            key, {'出勤': 0, '迟到': 0, '请假': 0, '早退': 0, '缺勤': 0, '总记录': 0})
        bucket[row['status']] += 1
        bucket['总记录'] += 1
    data = [[day, scene, item['出勤'], item['迟到'], item['请假'], item['早退'],
             item['缺勤'], item['总记录']]
            for (day, scene), item in sorted(grouped.items())]
    total = {'出勤': 0, '迟到': 0, '请假': 0, '早退': 0, '缺勤': 0, '总记录': 0}
    for item in grouped.values():
        for key in total:
            total[key] += item[key]
    data.append(['合计', '全部场景', total['出勤'], total['迟到'], total['请假'],
                 total['早退'], total['缺勤'], total['总记录']])
    return _sheet_bytes('考勤汇总', headers, data), '考勤汇总.xlsx'
