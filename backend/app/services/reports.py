# -*- coding: utf-8 -*-
"""周报、月报、成长报告和学期档案服务。

报告只负责组合已有业务服务的结果，并把生成时的结果保存为只读快照；
不在这里复制成绩、考勤或工作项的计算规则。
"""
from __future__ import annotations

from calendar import monthrange
from datetime import date, timedelta
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .. import clock, db
from . import attendance, class_context, comments, points, scores, work_items


REPORT_TYPES = {
    'weekly': '班级周报',
    'monthly': '班级月报',
    'term': '学期档案',
    'student_growth': '学生成长报告',
}


class ReportError(ValueError):
    pass


def _conn(conn=None):
    return conn or db.get_conn()


def _scope(conn=None):
    return class_context.get_current_scope(conn=_conn(conn))


def _date(value: str, field: str) -> str:
    text = str(value or '')[:10]
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise ReportError(f'{field}必须为 YYYY-MM-DD') from exc


def _period(report_type: str, period_start: str = '', period_end: str = '', conn=None):
    if report_type not in REPORT_TYPES:
        raise ReportError('报告类型不支持')
    today = clock.today()
    scope = _scope(conn)
    if period_start or period_end:
        start = _date(period_start, '开始日期')
        end = _date(period_end, '结束日期')
    elif report_type == 'weekly':
        start = (today - timedelta(days=today.weekday())).isoformat()
        end = (today + timedelta(days=6 - today.weekday())).isoformat()
    elif report_type == 'monthly':
        start = today.replace(day=1).isoformat()
        end = today.replace(day=monthrange(today.year, today.month)[1]).isoformat()
    else:
        start = str(scope.get('start_date') or f'{today.year}-01-01')[:10]
        end = str(scope.get('end_date') or today.isoformat())[:10]
        start = _date(start, '开始日期')
        end = _date(end, '结束日期')
    if start > end:
        raise ReportError('开始日期不能晚于结束日期')
    return start, end


def _active_students(conn, class_id, term_id):
    return [dict(row) for row in conn.execute(
        '''SELECT s.id, s.学号, s.姓名, s.性别
           FROM students s JOIN student_enrollments e ON e.student_id=s.id
           WHERE e.class_id=? AND e.term_id=? AND e.status='在读' AND s.deleted_at=''
           ORDER BY s.学号, s.id''', (class_id, term_id)
    ).fetchall()]


def _student(conn, student_id, class_id, term_id):
    row = conn.execute(
        '''SELECT s.id, s.学号, s.姓名, s.性别
           FROM students s JOIN student_enrollments e ON e.student_id=s.id
           WHERE s.id=? AND e.class_id=? AND e.term_id=? AND e.status='在读'
             AND s.deleted_at='' ''', (int(student_id), class_id, term_id)
    ).fetchone()
    if not row:
        raise ReportError('学生不存在或不在当前班级/学期')
    return dict(row)


def _education_sources(conn, class_id, term_id, start, end, student_id=None):
    params = [class_id, term_id, start, end]
    student_filter = ''
    if student_id is not None:
        student_filter = ' AND EXISTS (SELECT 1 FROM meeting_participants mp WHERE mp.meeting_id=m.id AND mp.student_id=?)'
        params.append(student_id)
    meetings = [dict(row) for row in conn.execute(
        '''SELECT m.id, m.held_on AS date, m.topic AS title, m.conclusion, '班会' AS kind
           FROM meeting_records m
           WHERE m.class_id=? AND m.term_id=? AND m.deleted_at='' AND m.held_on BETWEEN ? AND ?'''
        + student_filter + ' ORDER BY m.held_on, m.id', tuple(params)
    ).fetchall()]
    params = [class_id, term_id, start, end]
    if student_id is not None:
        params.append(student_id)
        student_filter = ' AND EXISTS (SELECT 1 FROM activity_participants ap WHERE ap.activity_id=a.id AND ap.student_id=?)'
    else:
        student_filter = ''
    activities = [dict(row) for row in conn.execute(
        '''SELECT a.id, a.occurred_on AS date, a.name AS title, a.result, a.retrospective, '活动' AS kind
           FROM activity_records a
           WHERE a.class_id=? AND a.term_id=? AND a.deleted_at='' AND a.occurred_on BETWEEN ? AND ?'''
        + student_filter + ' ORDER BY a.occurred_on, a.id', tuple(params)
    ).fetchall()]
    diaries = [dict(row) for row in conn.execute(
        '''SELECT d.id, d.diary_date AS date, d.work, d.event, d.reflection, d.todo, '日志' AS kind
           FROM diary_entries d
           WHERE d.class_id=? AND d.term_id=? AND d.deleted_at='' AND d.diary_date BETWEEN ? AND ?
           ORDER BY d.diary_date, d.id''', (class_id, term_id, start, end)
    ).fetchall()]
    return meetings, activities, diaries


def _events(conn, class_id, term_id, start, end, student_id=None):
    where = ['e.class_id=?', 'e.term_id=?', "e.deleted_at=''", 'substr(e.occurred_at,1,10) BETWEEN ? AND ?']
    params = [class_id, term_id, start, end]
    if student_id is not None:
        where.append('e.student_id=?')
        params.append(student_id)
    return [dict(row) for row in conn.execute(
        '''SELECT e.id, e.student_id, s.学号, s.姓名 AS student_name, e.occurred_at AS date,
                  e.event_type, e.description, e.status
           FROM student_events e JOIN students s ON s.id=e.student_id
           WHERE ''' + ' AND '.join(where) + ' ORDER BY e.occurred_at, e.id', tuple(params)
    ).fetchall()]


def _communications(conn, class_id, term_id, start, end, student_id=None):
    where = ['c.class_id=?', 'c.term_id=?', "c.deleted_at=''", 'substr(c.communicated_at,1,10) BETWEEN ? AND ?']
    params = [class_id, term_id, start, end]
    if student_id is not None:
        where.append('c.student_id=?')
        params.append(student_id)
    return [dict(row) for row in conn.execute(
        '''SELECT c.id, c.student_id, s.学号, s.姓名 AS student_name, c.communicated_at AS date,
                  c.method, c.reason, c.summary, c.status
           FROM communications c JOIN students s ON s.id=c.student_id
           WHERE ''' + ' AND '.join(where) + ' ORDER BY c.communicated_at, c.id', tuple(params)
    ).fetchall()]


def _summary(report_type, start, end, student_id=None, conn=None):
    conn = _conn(conn)
    scope = _scope(conn)
    class_id, term_id = scope['class_id'], scope['term_id']
    students = [_student(conn, student_id, class_id, term_id)] if student_id is not None else _active_students(conn, class_id, term_id)
    selected_ids = {int(item['id']) for item in students}
    attendance_rows = attendance.list_records(
        date_from=start, date_to=end, student_id=student_id, limit=50_000, conn=conn)
    attendance_counts = {status: 0 for status in attendance.STATUSES}
    for row in attendance_rows:
        attendance_counts[row['status']] = attendance_counts.get(row['status'], 0) + 1
    items = work_items.list_work_items(
        date_from=start, date_to=end, student_id=student_id, limit=50_000, conn=conn)
    point_rows = points.list_entries(student_id=student_id, date_from=start, date_to=end, limit=50_000, conn=conn)
    valid_points = [row for row in point_rows if row.get('status') == '有效']
    exam_rows = scores.list_records(student_id=student_id, conn=conn)
    exam_rows = [row for row in exam_rows if start <= str(row.get('exam_date') or '')[:10] <= end]
    comment_rows = comments.list_comments(student_id=student_id, limit=500, conn=conn)
    meetings, activities, diaries = _education_sources(conn, class_id, term_id, start, end, student_id)
    events = _events(conn, class_id, term_id, start, end, student_id)
    communications = _communications(conn, class_id, term_id, start, end, student_id)

    source_refs = {
        'attendance': [{'id': row['id'], 'date': row['attendance_date'], 'student_id': row['student_id'], 'status': row['status']} for row in attendance_rows],
        'work_items': [{'id': row['id'], 'title': row['title'], 'status': row['status'], 'source_type': row.get('source_type', '')} for row in items],
        'points': [{'id': row['id'], 'date': row.get('occurred_at', ''), 'student_id': row.get('student_id'), 'amount': row.get('amount', 0)} for row in point_rows],
        'scores': [{'id': row['id'], 'exam_name': row.get('exam_name', ''), 'exam_date': row.get('exam_date', ''), 'student_id': row.get('student_id')} for row in exam_rows],
        'comments': [{'id': row['id'], 'student_id': row.get('student_id'), 'status': row.get('status', '')} for row in comment_rows],
        'meetings': [{'id': row['id'], 'date': row['date'], 'title': row['title']} for row in meetings],
        'activities': [{'id': row['id'], 'date': row['date'], 'title': row['title']} for row in activities],
        'diary': [{'id': row['id'], 'date': row['date']} for row in diaries],
        'events': [{'id': row['id'], 'date': row['date'], 'student_id': row['student_id']} for row in events],
        'communications': [{'id': row['id'], 'date': row['date'], 'student_id': row['student_id']} for row in communications],
    }
    status_counts = {}
    for row in items:
        status_counts[row['status']] = status_counts.get(row['status'], 0) + 1
    report = {
        'report_type': report_type,
        'report_label': REPORT_TYPES[report_type],
        'period_start': start,
        'period_end': end,
        'scope': {'class_id': class_id, 'term_id': term_id, 'class_name': scope.get('class_name', ''), 'term_name': scope.get('term_name', '')},
        'student': students[0] if student_id is not None else None,
        'metrics': {
            'student_count': len(students),
            'attendance_total': len(attendance_rows),
            'attendance': attendance_counts,
            'work_items_total': len(items),
            'work_items_by_status': status_counts,
            'points_total': sum(float(row.get('amount') or 0) for row in valid_points),
            'points_entries': len(valid_points),
            'score_records': len(exam_rows),
            'comments': len(comment_rows),
            'events': len(events),
            'communications': len(communications),
            'meetings': len(meetings),
            'activities': len(activities),
            'diary_entries': len(diaries),
        },
        'sections': {
            'students': students,
            'attendance': attendance_rows,
            'work_items': items,
            'points': point_rows,
            'scores': exam_rows,
            'comments': comment_rows,
            'meetings': meetings,
            'activities': activities,
            'diary': diaries,
            'events': events,
            'communications': communications,
        },
        'source_refs': source_refs,
        'data_notes': [
            '成绩统计沿用结构化成绩服务的缺考、免考和未录入口径。',
            '积分只汇总有效流水，撤销记录保留在来源清单中。',
            '报告是生成时的只读快照，原始业务记录不会被修改。',
        ],
    }
    return report


def build_report(report_type: str, period_start: str = '', period_end: str = '', student_id: int | None = None, conn=None):
    conn = _conn(conn)
    start, end = _period(report_type, period_start, period_end, conn=conn)
    return _summary(report_type, start, end, student_id=student_id, conn=conn)


def create_archive(report_type: str, period_start: str = '', period_end: str = '', student_id: int | None = None, conn=None):
    conn = _conn(conn)
    report = build_report(report_type, period_start, period_end, student_id, conn=conn)
    scope = report['scope']
    title = report['report_label'] + (f' · {report["student"]["姓名"]}' if report.get('student') else '')
    row = conn.execute(
        '''INSERT INTO report_archives(class_id, term_id, report_type, period_start, period_end,
                                       student_id, title, payload_json)
           VALUES(?,?,?,?,?,?,?,?)
           ON CONFLICT(class_id, term_id, report_type, period_start, period_end, student_id)
           DO UPDATE SET title=excluded.title, payload_json=excluded.payload_json,
                         archived_at=datetime('now','localtime')
           RETURNING *''',
        (scope['class_id'], scope['term_id'], report_type, report['period_start'], report['period_end'],
         student_id, title, json.dumps(report, ensure_ascii=False)),
    ).fetchone()
    conn.commit()
    result = dict(row)
    result['payload'] = report
    return result


def list_archives(report_type: str = '', conn=None):
    conn = _conn(conn)
    class_id, term_id = class_context.scope_ids(conn=conn)
    where = ['class_id=?', 'term_id=?']
    params = [class_id, term_id]
    if report_type:
        if report_type not in REPORT_TYPES:
            raise ReportError('报告类型不支持')
        where.append('report_type=?')
        params.append(report_type)
    rows = [dict(row) for row in conn.execute(
        'SELECT id, report_type, period_start, period_end, student_id, title, created_at, archived_at '
        'FROM report_archives WHERE ' + ' AND '.join(where) + ' ORDER BY archived_at DESC, id DESC', tuple(params)
    ).fetchall()]
    return rows


def get_archive(archive_id: int, conn=None):
    conn = _conn(conn)
    class_id, term_id = class_context.scope_ids(conn=conn)
    row = conn.execute(
        'SELECT * FROM report_archives WHERE id=? AND class_id=? AND term_id=?',
        (int(archive_id), class_id, term_id),
    ).fetchone()
    if not row:
        raise ReportError('报告归档不存在')
    result = dict(row)
    result['payload'] = json.loads(result.pop('payload_json') or '{}')
    return result


def export_archive(archive_id: int, conn=None):
    report = get_archive(archive_id, conn=conn)
    payload = report['payload']
    wb = Workbook()
    ws = wb.active
    ws.title = '报告摘要'
    fill = PatternFill('solid', fgColor='5B6ABF')
    for index, (key, value) in enumerate([
        ('报告', report['title']), ('开始日期', report['period_start']), ('结束日期', report['period_end']),
        ('班级', payload.get('scope', {}).get('class_name', '')), ('学期', payload.get('scope', {}).get('term_name', '')),
    ], 1):
        ws.cell(index, 1, key).font = Font(bold=True, color='FFFFFF')
        ws.cell(index, 1).fill = fill
        ws.cell(index, 2, value)
    ws.cell(7, 1, '指标').font = Font(bold=True, color='FFFFFF')
    ws.cell(7, 1).fill = fill
    ws.cell(7, 2, '数值').font = Font(bold=True, color='FFFFFF')
    ws.cell(7, 2).fill = fill
    row_no = 8
    for key, value in payload.get('metrics', {}).items():
        ws.cell(row_no, 1, key)
        ws.cell(row_no, 2, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value)
        row_no += 1
    source_ws = wb.create_sheet('来源追溯')
    source_ws.append(['来源类型', '来源 ID', '日期/标题', '附加信息'])
    for cell in source_ws[1]:
        cell.fill = fill
        cell.font = Font(color='FFFFFF', bold=True)
    for kind, rows in payload.get('source_refs', {}).items():
        for item in rows:
            source_ws.append([kind, item.get('id', ''), item.get('date') or item.get('title') or '', json.dumps(item, ensure_ascii=False)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f'{report["title"]}_{report["period_start"]}_{report["period_end"]}.xlsx'
