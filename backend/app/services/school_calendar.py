# -*- coding: utf-8 -*-
"""学期校历：矩阵式 Excel 导入、日期查询和人工修正。"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from .. import clock, db
from . import audit, class_context


DAY_TYPES = {'上课日', '放假日', '调休上课', '考试日', '活动日', '其他'}
MONTHS = {
    '一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
    '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12,
}
WEEKS = {
    '一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7,
    '八': 8, '九': 9, '十': 10, '十一': 11, '十二': 12, '十三': 13,
    '十四': 14, '十五': 15, '十六': 16, '十七': 17, '十八': 18, '十九': 19,
    '二十': 20, '二十一': 21, '二十二': 22,
}
WEEKDAY_COLUMNS = ('星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日')
REST_WORDS = ('休', '放假', '假期', '春假', '寒假', '暑假', '清明', '劳动节', '端午', '国庆', '中秋')
EXAM_WORDS = ('考试', '高考', '中考', '合格考')
ACTIVITY_WORDS = ('报名', '典礼', '开学', '儿童节', '活动', '运动会', '培训')


class CalendarError(ValueError):
    pass


def _text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm(value) -> str:
    return re.sub(r'\s+', '', _text(value)).lower()


def _year_from_filename(filename: str, fallback: int) -> int:
    match = re.search(r'(20\d{2})', str(filename or ''))
    return int(match.group(1)) if match else fallback


def _month_number(value: str) -> int | None:
    text = _text(value).replace(' ', '')
    for name, number in MONTHS.items():
        if name in text:
            return number
    match = re.search(r'(?<!\d)(1[0-2]|[1-9])月', text)
    return int(match.group(1)) if match else None


def _week_number(value: str) -> int | None:
    text = _text(value).replace('周', '').replace('第', '').strip()
    if text.isdigit():
        return int(text)
    return WEEKS.get(text)


def _parse_date(value, *, default_year: int, default_month: int | None = None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return None
    match = re.search(r'(20\d{2})\s*[年/-]\s*(\d{1,2})\s*[月/-]\s*(\d{1,2})', text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.search(r'(?<!\d)(\d{1,2})\s*月\s*(\d{1,2})\s*日?', text)
    if match:
        return date(default_year, int(match.group(1)), int(match.group(2)))
    match = re.search(r'(?<!\d)(\d{1,2})\s*[/-]\s*(\d{1,2})(?!\d)', text)
    if match and default_month is None:
        return date(default_year, int(match.group(1)), int(match.group(2)))
    if default_month is not None:
        match = re.search(r'(?<!\d)(\d{1,2})(?!\d)', text)
        if match:
            return date(default_year, default_month, int(match.group(1)))
    return None


def _title_from_cell(value) -> str:
    text = _text(value)
    if not text:
        return ''
    titles = []
    for line in re.split(r'[\r\n]+', text):
        line = line.strip()
        if not line:
            continue
        line = re.sub(r'^(?:一月|二月|三月|四月|五月|六月|七月|八月|九月|十月|十一月|十二月|\d{1,2}月)/?\d{1,2}', '', line)
        line = re.sub(r'^\d{1,2}月', '', line)
        line = re.sub(r'^\d{1,2}(?=[^\d])', '', line)
        line = re.sub(r'(?<=[^\d])\d{1,2}$', '', line)
        line = re.sub(r'[（(]\s*休\s*[）)]', '', line)
        line = line.strip(' /-_')
        if re.fullmatch(r'\d{1,2}', line):
            continue
        if line and line not in titles:
            titles.append(line)
    return '、'.join(titles)


def _infer_day_type(raw: str, weekday: int, explicit: str = '') -> tuple[str, bool]:
    explicit = _text(explicit)
    if explicit in DAY_TYPES:
        return explicit, explicit in {'上课日', '调休上课', '考试日', '活动日'}
    text = _text(raw)
    if '调休' in text or '上班' in text or '补课' in text:
        return '调休上课', True
    if any(word in text for word in EXAM_WORDS):
        return '考试日', True
    if any(word in text for word in REST_WORDS):
        return '放假日', False
    if any(word in text for word in ACTIVITY_WORDS):
        return '活动日', True
    if weekday >= 5:
        return '放假日', False
    return '上课日', True


def _bool_value(value, fallback: bool) -> bool:
    text = _text(value).lower()
    if not text:
        return fallback
    if text in {'1', 'true', 'yes', '是', '上课', '工作日'}:
        return True
    if text in {'0', 'false', 'no', '否', '不上课', '休息日'}:
        return False
    return fallback


def _scope(*, write: bool = False, conn=None) -> tuple[int, int, dict]:
    conn = conn or db.get_conn()
    scope = class_context.get_current_scope(write=write, conn=conn)
    return int(scope['class_id']), int(scope['term_id']), scope


def _term_bounds(scope: dict) -> tuple[date | None, date | None]:
    def parse(value):
        try:
            return datetime.strptime(str(value or '')[:10], '%Y-%m-%d').date()
        except ValueError:
            return None
    return parse(scope.get('start_date')), parse(scope.get('end_date'))


def _entry_from_parts(calendar_date: date, raw: str, *, row: int, source: str = 'import',
                      explicit_type: str = '', explicit_school_day=None, note: str = '') -> dict:
    day_type, school_day = _infer_day_type(raw, calendar_date.weekday(), explicit_type)
    school_day = _bool_value(explicit_school_day, school_day)
    return {
        'row': row,
        'date': calendar_date.isoformat(),
        'day_type': day_type,
        'title': _title_from_cell(raw),
        'is_school_day': school_day,
        'note': _text(note),
        'source': source,
    }


def _matrix_rows(ws, filename: str, scope: dict) -> list[dict]:
    header = None
    for row in range(1, min(ws.max_row, 12) + 1):
        values = [_norm(ws.cell(row=row, column=col).value) for col in range(1, min(ws.max_column, 10) + 1)]
        if '月份' in values and '周次' in values and all(_norm(item) in values for item in WEEKDAY_COLUMNS):
            header = row
            break
    if not header:
        return []

    fallback_year = _year_from_filename(filename, clock.today().year)
    current_month = None
    first_week = None
    first_day = None
    first_col = None
    week_rows = []
    for row in range(header + 1, ws.max_row + 1):
        month = _month_number(_text(ws.cell(row=row, column=1).value))
        if month:
            current_month = month
        cells = [_text(ws.cell(row=row, column=col).value) for col in range(3, 10)]
        if not any(cells):
            continue
        week = _week_number(_text(ws.cell(row=row, column=2).value))
        if week is None:
            week = (week_rows[-1][0] + 1) if week_rows else None
        if week is None:
            continue
        week_rows.append((week, row, current_month, cells))
        if first_week is None:
            for index, cell in enumerate(cells):
                match = re.search(r'(?<!\d)(\d{1,2})(?!\d)', cell)
                if match and current_month:
                    first_week, first_day, first_col = week, int(match.group(1)), index
                    break
    if first_week is None or first_day is None or first_col is None:
        raise CalendarError('未能从校历矩阵识别首周日期，请检查月份、周次和星期列')

    first_date = date(fallback_year, week_rows[0][2] or 1, first_day)
    base_monday = first_date - timedelta(days=first_col)
    rows = []
    for week, row, _month, cells in week_rows:
        for index, raw in enumerate(cells):
            if not raw:
                continue
            calendar_date = base_monday + timedelta(days=(week - first_week) * 7 + index)
            rows.append(_entry_from_parts(calendar_date, raw, row=row))
    return rows


def _flat_rows(ws, filename: str) -> list[dict]:
    header = None
    mapping = {}
    aliases = {
        '日期': 'date', '日历日期': 'date', 'calendar_date': 'date', '开始日期': 'start_date',
        '结束日期': 'end_date', '类型': 'day_type', '日期类型': 'day_type', '安排类型': 'day_type',
        '事项': 'title', '安排': 'title', '内容': 'title', '名称': 'title', '备注': 'note',
        '是否上课': 'is_school_day', '是否行课': 'is_school_day',
    }
    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            key = _norm(ws.cell(row=row, column=col).value)
            if key in aliases:
                mapping[aliases[key]] = col
        if 'date' in mapping or 'start_date' in mapping:
            header = row
            break
    if not header:
        return []
    year = _year_from_filename(filename, clock.today().year)
    rows = []
    for row in range(header + 1, ws.max_row + 1):
        def value(key):
            col = mapping.get(key)
            return ws.cell(row=row, column=col).value if col else ''
        start = _parse_date(value('date') or value('start_date'), default_year=year)
        if not start:
            continue
        end = _parse_date(value('end_date'), default_year=year) or start
        if end < start or (end - start).days > 366:
            raise CalendarError(f'第 {row} 行的日期范围不合法')
        raw = f"{_text(value('title'))} {_text(value('day_type'))}".strip()
        for offset in range((end - start).days + 1):
            rows.append(_entry_from_parts(
                start + timedelta(days=offset), raw, row=row,
                explicit_type=_text(value('day_type')),
                explicit_school_day=value('is_school_day'), note=_text(value('note')),
            ))
    return rows


def _parse_workbook(file_bytes: bytes, filename: str, scope: dict) -> tuple[str, list[dict]]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise CalendarError(f'文件无法解析：{exc}') from exc
    try:
        for worksheet in workbook.worksheets:
            rows = _matrix_rows(worksheet, filename, scope)
            if rows:
                return 'matrix', rows
            rows = _flat_rows(worksheet, filename)
            if rows:
                return 'flat', rows
    finally:
        workbook.close()
    raise CalendarError('未识别到校历数据。支持“月份/周次/星期一至星期日”矩阵，或包含日期列的明细表。')


def preview_import(file_bytes: bytes, filename: str = '') -> dict:
    conn = db.get_conn()
    class_id, term_id, scope = _scope(conn=conn)
    result = {
        'filename': filename, 'format': '', 'rows': [], 'errors': [],
        'term': {'name': scope['term_name'], 'start_date': scope['start_date'], 'end_date': scope['end_date']},
        'summary': {'parsed': 0, 'valid': 0, 'new': 0, 'update': 0, 'skip': 0, 'conflict': 0, 'out_of_term': 0, 'error': 0},
    }
    try:
        format_name, parsed = _parse_workbook(file_bytes, filename, scope)
    except CalendarError as exc:
        result['errors'] = [{'row': 0, 'message': str(exc)}]
        result['summary']['error'] = 1
        return result
    result['format'] = format_name
    result['summary']['parsed'] = len(parsed)
    start, end = _term_bounds(scope)
    existing = {
        row['calendar_date']: dict(row) for row in conn.execute(
            '''SELECT * FROM school_calendar_days
               WHERE class_id=? AND term_id=? ORDER BY calendar_date''', (class_id, term_id)
        ).fetchall()
    }
    seen = {}
    for item in parsed:
        current = existing.get(item['date'])
        duplicate = seen.get(item['date'])
        out_of_term = bool((start and date.fromisoformat(item['date']) < start) or (end and date.fromisoformat(item['date']) > end))
        conflict = bool(duplicate and any(item[key] != duplicate[key] for key in ('day_type', 'title', 'is_school_day', 'note')))
        action = '新增' if not current else '更新'
        error = ''
        if conflict:
            action, error = '冲突', f'同一文件第 {duplicate["row"]} 行与第 {item["row"]} 行对同一天有不同安排'
            result['summary']['conflict'] += 1
        elif current and all(str(current.get(key) or '') == str(item[key] or '') for key in ('day_type', 'title', 'is_school_day', 'note')):
            action = '跳过'
            result['summary']['skip'] += 1
        elif current:
            result['summary']['update'] += 1
        else:
            result['summary']['new'] += 1
        if out_of_term:
            result['summary']['out_of_term'] += 1
        row = {**item, 'action': action, 'valid': not conflict, 'out_of_term': out_of_term, 'error': error}
        result['rows'].append(row)
        seen.setdefault(item['date'], item)
    result['summary']['valid'] = sum(1 for row in result['rows'] if row['valid'] and row['action'] != '跳过')
    return result


def _validate_entry(item: dict) -> dict:
    calendar_date = _parse_date(item.get('date') or item.get('calendar_date'), default_year=clock.today().year)
    if not calendar_date:
        raise CalendarError('校历日期不能为空且必须是 YYYY-MM-DD')
    day_type = _text(item.get('day_type')) or '上课日'
    if day_type not in DAY_TYPES:
        raise CalendarError(f'不支持的日期类型：{day_type}')
    return {
        'date': calendar_date.isoformat(), 'day_type': day_type,
        'title': _text(item.get('title')), 'is_school_day': int(_bool_value(item.get('is_school_day'), day_type != '放假日')),
        'note': _text(item.get('note')), 'row': int(item.get('row') or 0),
    }


def commit_import(rows: list[dict], filename: str = '', request_id: str = '') -> dict:
    conn = db.get_conn()
    class_id, term_id, _scope_data = _scope(write=True, conn=conn)
    request_id = _text(request_id) or f'calendar-{datetime.now().strftime("%Y%m%d%H%M%S%f")}'
    previous = conn.execute(
        '''SELECT imported, updated, skipped, conflict_count, error_count
           FROM school_calendar_import_runs WHERE class_id=? AND term_id=? AND request_id=?''',
        (class_id, term_id, request_id),
    ).fetchone()
    if previous:
        return {'idempotent': True, **dict(previous), 'errors': []}
    result = {'imported': 0, 'updated': 0, 'skipped': 0, 'conflict_count': 0, 'error_count': 0, 'errors': []}
    seen = set()
    try:
        for raw in rows or []:
            if not raw.get('valid', True) or raw.get('action') == '冲突':
                result['conflict_count'] += 1
                continue
            try:
                item = _validate_entry(raw)
            except CalendarError as exc:
                result['error_count'] += 1
                result['errors'].append({'row': raw.get('row', 0), 'message': str(exc)})
                continue
            if item['date'] in seen:
                result['conflict_count'] += 1
                continue
            seen.add(item['date'])
            current = conn.execute(
                '''SELECT id, day_type, title, is_school_day, note FROM school_calendar_days
                   WHERE class_id=? AND term_id=? AND calendar_date=?''',
                (class_id, term_id, item['date']),
            ).fetchone()
            if current and all(str(current[key] or '') == str(item[key] or '') for key in ('day_type', 'title', 'is_school_day', 'note')):
                result['skipped'] += 1
                continue
            if current:
                conn.execute(
                    '''UPDATE school_calendar_days SET day_type=?, title=?, is_school_day=?, note=?,
                       source='import', source_filename=?, source_row=?, updated_at=datetime('now','localtime')
                       WHERE id=?''',
                    (item['day_type'], item['title'], item['is_school_day'], item['note'], filename or '', item['row'], current['id']),
                )
                result['updated'] += 1
            else:
                conn.execute(
                    '''INSERT INTO school_calendar_days
                       (class_id, term_id, calendar_date, day_type, title, is_school_day, note, source, source_filename, source_row)
                       VALUES(?,?,?,?,?,?,?,?,?,?)''',
                    (class_id, term_id, item['date'], item['day_type'], item['title'], item['is_school_day'], item['note'], 'import', filename or '', item['row']),
                )
                result['imported'] += 1
        conn.execute(
            '''INSERT INTO school_calendar_import_runs
               (class_id, term_id, request_id, filename, imported, updated, skipped, conflict_count, error_count)
               VALUES(?,?,?,?,?,?,?,?,?)''',
            (class_id, term_id, request_id, filename or '', result['imported'], result['updated'], result['skipped'], result['conflict_count'], result['error_count']),
        )
        audit.record('school_calendar', request_id, 'import', summary='导入校历', params={
            'filename': filename or '', 'imported': result['imported'], 'updated': result['updated'],
            'skipped': result['skipped'], 'conflicts': result['conflict_count'],
        }, class_id=class_id, term_id=term_id, conn=conn, commit=False)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return result


def list_calendar(date_from: str = '', date_to: str = '', month: str = '') -> dict:
    conn = db.get_conn()
    class_id, term_id, scope = _scope(conn=conn)
    date_from = _text(date_from)[:10]
    date_to = _text(date_to)[:10]
    if month and re.fullmatch(r'20\d{2}-\d{2}', month):
        date_from = f'{month}-01'
        next_month = (datetime.strptime(date_from, '%Y-%m-%d').date().replace(day=28) + timedelta(days=4)).replace(day=1)
        date_to = (next_month - timedelta(days=1)).isoformat()
    where = ['class_id=?', 'term_id=?']
    params: list = [class_id, term_id]
    if date_from:
        where.append('calendar_date>=?')
        params.append(date_from)
    if date_to:
        where.append('calendar_date<=?')
        params.append(date_to)
    rows = [dict(row) for row in conn.execute(
        'SELECT * FROM school_calendar_days WHERE ' + ' AND '.join(where) + ' ORDER BY calendar_date', tuple(params)
    ).fetchall()]
    for row in rows:
        row['is_school_day'] = bool(row['is_school_day'])
    return {
        'scope': {'class_id': class_id, 'term_id': term_id, 'term_name': scope['term_name'], 'start_date': scope['start_date'], 'end_date': scope['end_date']},
        'entries': rows,
        'summary': {
            'total': len(rows),
            'school_days': sum(1 for row in rows if row['is_school_day']),
            'non_school_days': sum(1 for row in rows if not row['is_school_day']),
            'events': sum(1 for row in rows if row['title']),
        },
    }


def create_entry(calendar_date: str, day_type: str = '上课日', title: str = '', is_school_day: bool = True,
                 note: str = '') -> dict:
    return _save_entry(None, calendar_date, day_type, title, is_school_day, note)


def update_entry(entry_id: int, calendar_date: str, day_type: str = '上课日', title: str = '',
                 is_school_day: bool = True, note: str = '') -> dict:
    return _save_entry(int(entry_id), calendar_date, day_type, title, is_school_day, note)


def _save_entry(entry_id: int | None, calendar_date: str, day_type: str, title: str,
                is_school_day: bool, note: str) -> dict:
    conn = db.get_conn()
    class_id, term_id, _scope_data = _scope(write=True, conn=conn)
    item = _validate_entry({'date': calendar_date, 'day_type': day_type, 'title': title, 'is_school_day': is_school_day, 'note': note})
    try:
        if entry_id:
            current = conn.execute(
                'SELECT id FROM school_calendar_days WHERE id=? AND class_id=? AND term_id=?',
                (entry_id, class_id, term_id),
            ).fetchone()
            if not current:
                raise CalendarError('校历记录不存在')
            conn.execute(
                '''UPDATE school_calendar_days SET calendar_date=?, day_type=?, title=?, is_school_day=?, note=?,
                   source='manual', source_filename='', source_row=NULL, updated_at=datetime('now','localtime') WHERE id=?''',
                (item['date'], item['day_type'], item['title'], item['is_school_day'], item['note'], entry_id),
            )
            result_id = entry_id
        else:
            result_id = conn.execute(
                '''INSERT INTO school_calendar_days(class_id, term_id, calendar_date, day_type, title, is_school_day, note, source)
                   VALUES(?,?,?,?,?,?,?, 'manual')''',
                (class_id, term_id, item['date'], item['day_type'], item['title'], item['is_school_day'], item['note']),
            ).lastrowid
        audit.record('school_calendar', result_id, 'update' if entry_id else 'create', summary='修改校历日期', params={
            'date': item['date'], 'day_type': item['day_type'], 'title': item['title'], 'is_school_day': bool(item['is_school_day']),
        }, class_id=class_id, term_id=term_id, conn=conn, commit=False)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return {'id': result_id}


def query_calendar(date_from: str = '', date_to: str = '', day_type: str = '', limit: int = 100) -> dict:
    data = list_calendar(date_from=date_from, date_to=date_to)
    entries = data['entries']
    if day_type:
        entries = [row for row in entries if row['day_type'] == str(day_type).strip()]
    limit = max(1, min(int(limit), 200))
    return {
        'date_from': _text(date_from)[:10], 'date_to': _text(date_to)[:10],
        'day_type': _text(day_type), 'entries': entries[:limit], 'count': len(entries[:limit]),
        'total_count': len(entries), 'truncated': len(entries) > limit,
    }
