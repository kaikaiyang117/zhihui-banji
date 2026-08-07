# -*- coding: utf-8 -*-
"""学生信息 Excel 导入：模板下载 + 文件解析 + 按学号合并去重"""
from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import Workbook, load_workbook

from .config import STUDENT_COLUMNS
from .db import get_conn
from .services import class_context

NORMALIZE = {
    '学号': '学号', '姓名': '姓名', '性别': '性别', '出生年月': '出生年月',
    '民族': '民族', '家庭住址': '家庭住址', '监护人姓名': '监护人姓名',
    '监护人电话': '监护人电话', '监护人职业': '监护人职业', '是否住校': '是否住校',
    '特长': '特长', '班级任职': '班级任职', '备注': '备注',
    '监护人2姓名': '监护人2姓名', '监护人2电话': '监护人2电话', '监护人2关系': '监护人2关系',
}


def _norm(s: str) -> str:
    return re.sub(r'\s+', '', str(s or '')).strip()


def _cell_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def build_template() -> io.BytesIO:
    """生成学生信息导入模板 xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = '学生信息'
    for i, h in enumerate(STUDENT_COLUMNS, 1):
        ws.cell(row=1, column=i, value=h)
        ws.cell(row=1, column=i).font = ws.cell(row=1, column=i).font.copy(weight='bold')
    for i, example in enumerate(['2201', '张三', '男', '2010-05', '汉', '汶川县威州镇',
                                 '张大明', '13800000000', '务农', '住校', '书法', '纪律委员', '',
                                 '李芳', '13900000000', '母亲'], 1):
        ws.cell(row=2, column=i, value=example)
    for col, width in zip('ABCDEFGHIJKLMNOP', [10, 10, 6, 10, 6, 22, 10, 14, 8, 8, 10, 10, 30, 10, 14, 10]):
        ws.column_dimensions[col].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _empty_summary() -> dict:
    return {'imported': 0, 'updated': 0, 'skipped': 0, 'valid': 0}


def preview_students(file_bytes: bytes, filename: str = '') -> dict:
    """只解析和校验学生 Excel，不修改数据库，返回待确认的有效行。"""
    conn = get_conn()
    result = {'filename': filename, 'rows': [], 'errors': [], 'summary': _empty_summary()}
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result['errors'] = [{'row': 0, 'msg': f'文件无法解析: {e}'}]
        result['summary']['skipped'] = 1
        return result

    ws = wb.active if wb.active is not None else wb.worksheets[0]

    # 定位表头行（第一个包含「学号」「姓名」的行）
    header_row = -1
    col_map = {}
    for r in range(1, min(ws.max_row, 10) + 1):
        row_hdr = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            key = _norm(v)
            if not key:
                continue
            mapped = NORMALIZE.get(key)
            if mapped:
                row_hdr[mapped] = c
        if '学号' in row_hdr and '姓名' in row_hdr:
            header_row = r
            col_map = row_hdr
            break

    if header_row <= 0:
        result['errors'] = [{'row': 1, 'msg': '未找到表头，需包含「学号」「姓名」列（列名不能带多余空格/换行）'}]
        result['summary']['skipped'] = 1
        return result

    def get_val(r, key):
        c = col_map.get(key)
        if not c:
            return ''
        return _cell_text(ws.cell(row=r, column=c).value)

    seen = set()
    for r in range(header_row + 1, ws.max_row + 1):
        xh = _norm(get_val(r, '学号'))
        name = _norm(get_val(r, '姓名'))
        if not xh and not name:
            continue  # 跳过空行
        if not xh:
            result['errors'].append({'row': r, 'msg': f'姓名「{name}」缺少学号，已跳过'})
            result['summary']['skipped'] += 1
            continue
        if not name:
            result['errors'].append({'row': r, 'msg': f'学号「{xh}」缺少姓名，已跳过'})
            result['summary']['skipped'] += 1
            continue
        if xh in seen:
            result['errors'].append({'row': r, 'msg': f'学号「{xh}」在文件中重复，已跳过'})
            result['summary']['skipped'] += 1
            continue
        fields = {k: get_val(r, k) for k in STUDENT_COLUMNS}
        fields['学号'] = xh
        fields['姓名'] = name
        existing = conn.execute('SELECT id, deleted_at FROM students WHERE 学号=?', (xh,)).fetchone()
        if existing and existing['deleted_at']:
            result['errors'].append({'row': r, 'msg': f'学号「{xh}」位于回收站，请先恢复'})
            result['summary']['skipped'] += 1
            continue
        action = '更新' if existing else '新增'
        result['rows'].append({
            'row': r,
            'action': action,
            'student_id': existing['id'] if existing else None,
            'fields': fields,
        })
        result['summary']["updated" if existing else "imported"] += 1
        seen.add(xh)

    result['summary']['valid'] = len(result['rows'])
    wb.close()
    return result


def commit_student_import(rows: list[dict], filename: str = '') -> dict:
    """提交预览后的有效行，提交前再次校验学号和文件内重复。"""
    conn = get_conn()
    class_id, term_id = class_context.scope_ids(write=True, conn=conn)
    result = {'imported': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    valid = []
    seen = set()
    for item in rows or []:
        fields = item.get('fields', {}) if isinstance(item, dict) else {}
        xh = _norm(fields.get('学号', ''))
        name = _norm(fields.get('姓名', ''))
        row_no = item.get('row', 0) if isinstance(item, dict) else 0
        if not xh or not name:
            result['errors'].append({'row': row_no, 'msg': '学号和姓名不能为空，已跳过'})
            result['skipped'] += 1
            continue
        if xh in seen:
            result['errors'].append({'row': row_no, 'msg': f'学号「{xh}」重复，已跳过'})
            result['skipped'] += 1
            continue
        normalized = {key: _cell_text(fields.get(key, '')) for key in STUDENT_COLUMNS}
        normalized['学号'] = xh
        normalized['姓名'] = name
        valid.append((row_no, normalized))
        seen.add(xh)

    try:
        for _, fields in valid:
            vals = tuple(fields.values())
            cols = ','.join(f'[{k}]' for k in STUDENT_COLUMNS)
            placeholders = ','.join('?' for _ in STUDENT_COLUMNS)
            existing = conn.execute('SELECT id, deleted_at FROM students WHERE 学号=?', (fields['学号'],)).fetchone()
            if existing and existing['deleted_at']:
                result['errors'].append({'row': _, 'msg': f'学号「{fields["学号"]}」位于回收站，请先恢复'})
                result['skipped'] += 1
                continue
            if existing:
                updates = ','.join(f'[{k}]=?' for k in STUDENT_COLUMNS)
                conn.execute(
                    f'UPDATE students SET {updates}, updated_at=datetime(\'now\',\'localtime\') WHERE 学号=?',
                    (*vals, fields['学号']))
                result['updated'] += 1
                student_id = existing['id']
            else:
                student_id = conn.execute(
                    f'INSERT INTO students({cols}) VALUES({placeholders})', vals)
                student_id = student_id.lastrowid
                result['imported'] += 1
            class_context.enroll_student(
                student_id, class_id, term_id, conn=conn, commit=False)
        conn.execute(
            '''INSERT INTO student_import_runs
               (filename, imported, updated, skipped, error_count, class_id, term_id)
               VALUES(?,?,?,?,?,?,?)''',
            (filename or '', result['imported'], result['updated'],
             result['skipped'], len(result['errors']), class_id, term_id))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    return result


def import_students(file_bytes: bytes, filename: str = '') -> dict:
    """兼容旧接口：解析并立即提交，不再作为前端默认流程。"""
    preview = preview_students(file_bytes, filename)
    result = commit_student_import(preview['rows'], filename)
    result['skipped'] += preview['summary']['skipped']
    result['errors'] = preview['errors'] + result['errors']
    return result
